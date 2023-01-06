import sys
from string import ascii_lowercase
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from sales_send_mail_reusable_task import send_mail, send_mail_with_attachment
# import send mail program

class average_day_sales_Exception(Exception):
    pass

def reading_main_config_sheet(path, sheet_name):
    try:
        dict_config_main = dict()
        work_book = openpyxl.load_workbook(path)
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        # print(maximum_row)
        maximum_col = work_sheet.max_column
        # print(maximum_col)
        for row in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = row[0].value
            value = row[1].value
            dict_config_main[key] = value
        work_book.close()
        return dict_config_main
    except Exception as config_error:
        print("failed to load main config file. Hence, stopping the BOT")
        print(config_error)
        sys.exit()


def leap_year(year):
    if type(year) != int:
        print("year value {0} passed to the function is not a number, hence stopping the bot")
        sys.exit()
    # divided by 100 means century year (ending with 00)
    # century year divided by 400 is leap year
    if (year % 400 == 0) and (year % 100 == 0):
        print("{0} is a leap year".format(year))
        bool_leap_year = True
    # not divided by 100 means not a century year
    #  divided by 4 is a leap year
    elif (year % 4 == 0) and (year % 100 != 0):
        print("{0} is a leap year".format(year))
        bool_leap_year = True
    # if not divided by both 400 (century year) and 4 (not century year)
    # year is not leap year
    else:
        print("{0} is not a leap year".format(year))
        bool_leap_year = False

    return bool_leap_year


def average_day_sales(str_sales_register_file_path, str_sales_register_file_sheet_name, main_config):

    skip_rows = main_config['sales_register_file_header_row_number'] - 1
    try:
        sales_register_df = pd.read_excel(str_sales_register_file_path, skiprows=skip_rows, sheet_name=str_sales_register_file_sheet_name)
        print(sales_register_df)
    except Exception as sales_register_file_read_exception:
        exception_message = "Below exception occurred while reading sales register file:"
        print(sales_register_file_read_exception)
        raise average_day_sales_Exception(exception_message)



    # creating a pivot table
    sales_register_columns_list = sales_register_df.columns.values.tolist()
    columns_not_found_list = []
    for column in ["Billing Date", "Base Price in INR"]:
        if column not in sales_register_columns_list:
            columns_not_found_list.append(column)
    if len(columns_not_found_list) != 0:
        print("below columns are missing in input sales register file, hence stopping the program. \n\t {0}".format(columns_not_found_list))
        # send mail notification
        sys.exit()
    try:
        average_day_sales_pivot_df = pd.pivot_table(sales_register_df, index="Billing Date", values="Base Price in INR", aggfunc=np.sum)
        print(average_day_sales_pivot_df)
    except Exception as pivot_table_exception:
        exception_message = "Below exception occurred while creating the pivot table for average day sales report,Hence stopping the program.\n\t {0}".format(pivot_table_exception)
        raise average_day_sales_Exception(exception_message)

        # reset index for having proper index for the dataframe and unsetting index column making usual column
    average_day_sales_pivot_df = average_day_sales_pivot_df.reset_index()
    print(average_day_sales_pivot_df)
    try:
        float_base_price_column_sum = average_day_sales_pivot_df["Base Price in INR"].sum()
        print("float_base_price_column_sum:", float_base_price_column_sum)
    except Exception as base_price_sum_exception:
        exception_message = "Exception occurred while calculating sum of values in base price column"
        print(base_price_sum_exception)
        raise average_day_sales_Exception(exception_message)

    try:
        average_day_sales_pivot_df["month"] = average_day_sales_pivot_df["Billing Date"].dt.month_name().str[:3]
    except Exception as month_column_finding_exception:
        exception_message = "Exception occurred while creating month column from billing date that ay be due to values other than dates"
        print(month_column_finding_exception)
        raise average_day_sales_Exception(exception_message)

    month_unique_values_list = average_day_sales_pivot_df["month"].unique().tolist()
    print(month_unique_values_list)
    try:
        average_day_sales_pivot_df["year"] = average_day_sales_pivot_df["Billing Date"].dt.year
    except Exception as year_column_finding_exception:
        exception_message = "Below Exception occurred while creating year column from billing date that ay be due to values other than dates. \n \t {0}".format(year_column_finding_exception)
        raise average_day_sales_Exception(exception_message)

    year_unique_values_list = average_day_sales_pivot_df["year"].unique().tolist()
    # print(year_unique_values_list)
    year = year_unique_values_list[0]
    if len(year_unique_values_list) != 1:
        print("sales register contains data of more than 1 year. hence stopping the bot")
        # send mail notification
        sys.exit()
    print(year)
    quarter = None
    q1_months = ['Apr', 'May', 'Jun']
    q2_months = ['Jul', 'Aug', 'Sep']
    q3_months = ['Oct', 'Nov', 'Dec']
    q4_months = ['Jan', 'Feb', 'Mar']

    if len(month_unique_values_list) == 3:
        for month in q1_months:
            if month in month_unique_values_list:
             quarter = q1_months
        for month in q2_months:
            if month in month_unique_values_list:
             quarter = q2_months
        for month in q3_months:
            if month in month_unique_values_list:
              quarter = q3_months

        for month in q4_months:
            if month in month_unique_values_list:
              quarter = q4_months
        print(quarter)
    else:
        print("sales register contains data of more than 3 months for a quarter. hence stopping the bot")
        # send mail notification
        sys.exit()

    if quarter == q4_months:
        # find year whether it is leap or not
        bool_leap_year=leap_year(year)
        if bool_leap_year:
            int_days = 91
        else:
            int_days = 90
    elif quarter == q1_months:
        int_days = 91
    elif quarter == q2_months or quarter == q3_months:
        int_days = 92
    else:
        print("The program could not find financial timeline information such as quarter, months and year of sales register")
        # send mail notification
        sys.exit()

    print(int_days)
    float_average_sales_for_day = float_base_price_column_sum / int_days
    print("float_average_sales_for_day:", float_average_sales_for_day)


    average_day_sales_pivot_df["average sales for day"] = float_average_sales_for_day
    try:
        float_average_sales_for_day_sum = average_day_sales_pivot_df["average sales for day"].sum()
        print(float_average_sales_for_day_sum)
    except Exception as average_sales_per_day_calculation_exception:
        exception_message = "Below exception occurred while calculating sum of 'average sales for day' column \n\t {0}".format(average_sales_per_day_calculation_exception)
        raise average_day_sales_Exception(exception_message)

    average_day_sales_pivot_df["Variance"] = average_day_sales_pivot_df["Base Price in INR"] - float_average_sales_for_day   # average
    try:
        float_variance_sum = average_day_sales_pivot_df["Variance"].sum()
        print(float_variance_sum)
    except Exception as variance_sum_exception:
        exception_message = "Below Exception occurred while creating year column from billing date that ay be due to values other than dates. \n \t {0}".format(variance_sum_exception)
        raise average_day_sales_Exception(exception_message)

    average_day_sales_pivot_df["Concentration"] = average_day_sales_pivot_df["Variance"]/float_variance_sum
    print(average_day_sales_pivot_df)

    # sorting the Concentration in descending order
    average_day_sales_pivot_df = average_day_sales_pivot_df.sort_values(by=["Concentration"], ascending=False)  # descending order
    # creating a column as remarks

    # And for those concentrations greater than 25% and -25% name them as major in remarks column
    positive_threshold_percentage = main_config['threshold_percentage']/100
    negative_threshold_percentage = main_config['threshold_percentage']/ (-100)
    remarks = main_config['remarks']
    average_day_sales_pivot_df["Remarks"] = ''
    for index, row in average_day_sales_pivot_df.iterrows():
        print(row['Concentration'])
        if row['Concentration'] >= positive_threshold_percentage or row['Concentration'] <= negative_threshold_percentage:
            average_day_sales_pivot_df.at[index, 'Remarks'] = remarks


    average_day_sales_pivot_df.drop("month", inplace=True, axis=1)
    average_day_sales_pivot_df.drop("year", inplace=True, axis=1)
    #  date time
    average_day_sales_pivot_df["Billing Date"] = pd.to_datetime(average_day_sales_pivot_df["Billing Date"]).dt.strftime("%d-%b")


    output_file_path = main_config['Output_file_path']
    output_sheet_name = main_config['average_day_sales_output_sheet_name']
    try:
        average_day_sales_pivot_df.to_excel(output_file_path, sheet_name=output_sheet_name, startrow=2, index=False)
    except Exception as file_saving_exception:
        exception_message = file_saving_exception
        raise average_day_sales_Exception(exception_message)

    # Load Sheet in openpyxl
    workbook = openpyxl.load_workbook(output_file_path)
    worksheet = workbook[output_sheet_name]
    # worksheet = workbook.active
    for row in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            print(col[row].value)
    print("workbook", workbook)
    worksheet['B2'] = float_base_price_column_sum  # "=sum(B6:B93)
    worksheet['C2'] = float_average_sales_for_day_sum  # sum of average  sales for day
    worksheet['D2'] = float_variance_sum

    # Number format implementation
    for cell in worksheet['B']:
       cell.number_format = "#,###,##"
    for cell in worksheet['C']:
       cell.number_format = "#,###,##"
    for cell in worksheet['D']:
       cell.number_format = "#,###,##"

    # Format Variance
    for cell in worksheet['E']:
        cell.number_format = '0.0%'

    # Format Header
    calibri_11_black_bold_font = Font(name="Calibri", size=11, color="000000", bold=True)
    print(ascii_lowercase)
    for c in ascii_lowercase:
        print(c)
        worksheet[c + "3"].font = calibri_11_black_bold_font

    # Header Fill
    solid_yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
    for f in ascii_lowercase:
       worksheet[f + "3"].fill = solid_yellow_fill
       if f == 'f':
          break

    thin = Side(border_style="thin", color='000000')
    for row in worksheet.iter_rows(min_row=3, min_col=1, max_row=worksheet.max_row, max_col=6):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Set Width
    for c in ascii_lowercase:
        worksheet.column_dimensions[c].width = 25

    workbook.save(output_file_path)
    return output_file_path
#  subject = main_config["ColumnMiss_Subject"]
#         body = main_config["ColumnMiss_Body"]
#         body = body.replace("ColumnName +")
#         send_mail(to=to_address, cc=cc, subject=subject, body=body)
#         sys.exit()

if __name__ == '__main__':
    str_path = "sales register q1.xlsx"
    str_sales_register_file_sheet_name = "Sheet1"

    config_file_path = 'Config_file.xlsx'
    config_file_sheet_name = 'Main'
    main_config = reading_main_config_sheet(path=config_file_path, sheet_name=config_file_sheet_name)
    try:
        to = main_config["to"]
        cc = main_config["cc"]
        average_day_sales_output_file_path = average_day_sales(str_sales_register_file_path=str_path,
                      str_sales_register_file_sheet_name=str_sales_register_file_sheet_name,
                      main_config=main_config)
        subject = "Average day sales report is successfully generated"
        body = '''
        Hello,
        
        Program successfully completed.
        
        Thanks,
        RPA Team
        '''


        send_mail_with_attachment(to=main_config["to"], cc=main_config["cc"], subject=subject, body=body, attachment_path=average_day_sales_output_file_path)
    except (KeyError, ValueError, TypeError, OSError, ImportError, MemoryError, RuntimeError, Exception) as average_day_sales_exception:
        to = main_config["to"]
        cc = main_config["cc"]

        subject = "Sales Register - Average Day sales Register sheet - Error Occurred"
        body ='''
        Hello,
        
        {0}
        
        Thanks,
        RPA Team
        
        '''.format(average_day_sales_exception)
        send_mail(to=to, cc=cc, subject=subject, body=body)
        print("Below exception occurred while creating the average day report sales report. \n\t {0}".format(average_day_sales_exception))
