import os

import openpyxl


class Config_Exception(Exception):
    pass


def create_config_dict_from_Config_file(path, sheet_name):
    try:

        dict_config_main = {}

        if not os.path.exists(path):
            raise Config_Exception("Exception: Config file is not exist in the path provided {0}".format(path))
        try:
            print("Config file path is exist")
            workbook = openpyxl.load_workbook(path)
            print("Config file is loaded")
        except Exception as config_file_error:
            exception_message = "Exception is occurred while loading Config file from path {0} because {1}".format(path,                                                                                                       config_file_error)
            raise Config_Exception(exception_message)

        try:
            worksheet = workbook[sheet_name]
            print("Config sheet is read")
        except Exception as work_sheet_exception:
            exception_message = "Exception is occurred while loading Config Excel file sheet {0} because {1}".format(
                sheet_name, work_sheet_exception)
            raise Config_Exception(exception_message)

        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for config_details in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = config_details[0].value
            value = config_details[1].value
            dict_config_main[key] = value
        print("Config dictionary is created from config file")
        try:
            workbook.close()
        except Exception as config_save_exception:
            exception_message = "Exception occurred while saving Config file in path {0} because {1}".format(path, config_save_exception)
            raise Config_Exception(exception_message)

        return dict_config_main

    except Exception as config_file_read_error:
        raise Config_Exception(config_file_read_error)




if __name__ =='__main__':
    path = 'Config_file.xlsx'
    sheet_name = 'Main'
    try:
        config_dict = create_config_dict_from_Config_file(path, sheet_name)
        print(config_dict)
    except Exception as config_dict_creation_exception:
        print(config_dict_creation_exception)
    # print(type(config_dict))
    # print(config_dict)
