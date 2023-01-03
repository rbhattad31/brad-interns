
import openpyxl

path = "demo_excel.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

print(sheet_obj.max_column)
print(sheet_obj.max_row)

