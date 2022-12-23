import openpyxl
path = "Excel.xlsx"
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active
print(worksheet)
cell_obj = worksheet.cell(row=1, column=1)
cell_obj1 = worksheet.cell(row=1, column=2)
cell_obj2 = worksheet.cell(row=1, column=3)
cell_obj3 = worksheet.cell(row=2, column=1)
cell_obj4 = worksheet.cell(row=2, column=2)
cell_obj5 = worksheet.cell(row=2, column=3)
print(cell_obj.value)
print(cell_obj1.value)
print(cell_obj2.value)
print(cell_obj3.value)
print(cell_obj4.value)
print(cell_obj5.value)
