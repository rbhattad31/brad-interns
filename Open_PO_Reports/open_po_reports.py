import os
import logging
import datetime
from string import ascii_lowercase
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Side, Border

from sales_send_mail_reusable_task import send_mail_with_attachment, send_mail


class open_po_Exception(Exception):
    pass


def create_config_dict_from_Config_file(path, sheet_name):
    try:

        dict_config_main = {}

        if not os.path.exists(path):
            raise open_po_Exception("Config file is not exist in the path provided")
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as config_file_error:
            exception_message = "Below error is occurred while loading Config file from path {0}. \n\t {1}".format(path, config_file_error)
            logging.error("Exception occurred while loading Config file from path")

            raise open_po_Exception(exception_message)
        try:
            worksheet = workbook[sheet_name]
            print(type(worksheet))
        except Exception as work_sheet_exception:
            exception_message = "Below error is occurred while loading Config Excel file sheet {0}.  \n\t {1}".format(sheet_name, work_sheet_exception)
            logging.error("Exception occurred while loading Config Excel file sheet")
            raise open_po_Exception(exception_message)

        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for config_details in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = config_details[0].value
            value = config_details[1].value
            dict_config_main[key] = value

        try:
            workbook.save(path)
        except Exception as config_save_exception:
            exception_message = "Below error is occurred while saving Config file in path {0} \n\t {1}".format(path, config_save_exception)
            logging.error("Exception occurred while saving Config file from path")
            raise open_po_Exception(exception_message)
        return dict_config_main

    except Exception as config_file_read_error:
        exception_message = "Below error is occurred while reading Config file in path {0} \n\t {1}".format(path, config_file_read_error)
        logging.error("Exception occurred while reading Config file in path")
        raise open_po_Exception(exception_message)


def open_po_reports(open_po_df, main_config):
    print(open_po_df)
    print(list(open_po_df.columns.values.tolist()))

    try:
        open_po_df["Order Date"] = pd.to_datetime(open_po_df["Order Date"], errors='coerce')
        open_po_df["PO Date"] = pd.to_datetime(open_po_df["PO Date"], errors='coerce')
    except Exception as changing_date_format_exception:
        exception_message = changing_date_format_exception
        logging.error("Exception occurred while changing the date format")
        raise open_po_Exception(exception_message)
    try:
        recent_audit_date = open_po_df["Order Date"].max()  # error
        print(recent_audit_date)
    except Exception as recent_audit_date_exception:
        exception_message = recent_audit_date_exception
        logging.error("Exception occurred while recent audit date exception the date format")
        raise open_po_Exception(exception_message)
    open_po_df["no of days"] = 0
    output_file_path = main_config['Output_file_path']
    print(output_file_path)
    output_sheet_name = main_config['open_po_report_output_sheet_name']
    print(output_sheet_name)

    for index in open_po_df.index:
        po_date = open_po_df.loc[index]["PO Date"]
        no_of_days = recent_audit_date - po_date
        # print(no_of_days)
        open_po_df.at[index, 'no of days'] = no_of_days

    try:
        open_po_df["PO Date"] = pd.to_datetime(open_po_df["PO Date"]).dt.strftime("%d-%m-%Y")
        open_po_df["Order Date"] = pd.to_datetime(open_po_df["Order Date"]).dt.strftime("%d-%m-%Y")
        print(open_po_df)
    except Exception as changing_date_format_exception:
        exception_message = changing_date_format_exception
        logging.error("Below exception occurred while changing the date format")
        raise open_po_Exception(exception_message)


    try:
        print(open_po_df)
        print(open_po_df.columns.values.tolist())
        open_po_df.to_excel(output_file_path, sheet_name=output_sheet_name, startrow=1, index=False)
    except Exception as file_saving_exception:
        exception_message = file_saving_exception
        logging.error("Below exception occurred while saving the file ")
        raise open_po_Exception(exception_message)

#   load in openpyxl
    workbook = openpyxl.load_workbook(output_file_path)
    worksheet = workbook[output_sheet_name]
    worksheet['P1'] = recent_audit_date  # date format change
    print(recent_audit_date)
    # value = datetime.datetime.strptime(recent_audit_date, "%Y-%m-%d %H:%M:%S")
    # cell = worksheet['P1']
    # cell.value = value
    # cell.number_format = '%d-%m-%Y'
        # Format Header
    calibri_11_black_bold_font = Font(name="Calibri", size=11, color="000000", bold=True)
    print(ascii_lowercase)
    for c in ascii_lowercase:
        print(c)
        worksheet[c + "2"].font = calibri_11_black_bold_font

    # Header Fill
    light_blue_fill = PatternFill(patternType='solid', fgColor='87ceeb')
    for c in ascii_lowercase:
        worksheet[c + "2"].fill = light_blue_fill
        if c == 'p':
            break

    thin = Side(border_style="thin", color='000000')
    for row in worksheet.iter_rows(min_row=3, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Set Width
    # for c in ascii_lowercase:
        # worksheet.column_dimensions[c].width = 30

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(output_file_path)
    return output_file_path


if __name__ == '__main__':
    log_file_path = 'default_log.iog'
    logging_level = 10

    logging.basicConfig(filename=log_file_path, level=logging_level, format='%(asctime)s::%(levelname)s::%(message)s')
    logging.info("testing logging")
    str_open_po_file_path = "open PO.xlsx"
    str_open_po_file_sheet_name = "OPEN PO"
    config_file_path = 'Config file open po.xlsx'
    config_file_sheet_name = 'open_po_report'

    try:
        dict_main_config = create_config_dict_from_Config_file(path=config_file_path, sheet_name=config_file_sheet_name)
        skip_rows = dict_main_config['open_po_file_header_row_number']
    except Exception as config_main_creation_exception:
        exception_message = "config main creation"
        logging.error("Below exception occurred while config main creation")
        raise open_po_Exception(exception_message)

    try:
        open_po_df = pd.read_excel(str_open_po_file_path, skiprows=skip_rows, sheet_name=str_open_po_file_sheet_name)
        print(skip_rows)
        print(open_po_df)
        print(open_po_df.columns.values.tolist())

    except Exception as open_po_file_read_exception:
        exception_message = open_po_file_read_exception
        logging.error("Exception occurred while reading sales register file")
        raise open_po_Exception(exception_message)

    try:
        to = dict_main_config["to"]
        cc = dict_main_config["cc"]
        open_po_reports_output_file_path = open_po_reports(open_po_df=open_po_df, main_config=dict_main_config)
        subject = "Open po report is successfully generated"
        body = '''
        Hello,

        Program successfully completed.

        Thanks,
        RPA Team
        '''

        send_mail_with_attachment(to=dict_main_config["to"], cc=dict_main_config["cc"], subject=subject, body=body,
                                  attachment_path=open_po_reports_output_file_path)
    except (KeyError, ValueError, TypeError, OSError, ImportError, MemoryError, RuntimeError,
            Exception) as open_po_Exception:
        to = dict_main_config["to"]
        cc = dict_main_config["cc"]

        subject = "open po - open po reports sheet - Error Occurred"
        body = '''
        Hello,

        {0}

        Thanks,
        RPA Team

        '''.format(open_po_Exception)
        send_mail(to=to, cc=cc, subject=subject, body=body)
        print("Below exception occurred while creating the open day report sales report. \n\t {0}".format(open_po_Exception))
