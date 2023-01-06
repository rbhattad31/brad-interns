import os

import openpyxl


def create_config_dict_from_Config_file(path, sheet_name):
    try:

        dict_config_main = {}

        if not os.path.exists(path):
            raise "Config file is not exist in the path provided"
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as config_file_error:
            exception_message = "Below error is occurred while loading Config file from path {0}. \n\t {1}".format(path,
                                                                                                                   config_file_error)
            raise exception_message

        try:
            worksheet = workbook[sheet_name]
        except Exception as work_sheet_exception:
            exception_message = "Below error is occurred while loading Config Excel file sheet {0}. \n\t {1}".format(
                sheet_name, work_sheet_exception)
            raise exception_message

        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for config_details in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = config_details[0].value
            value = config_details[1].value
            dict_config_main[key] = value

        try:
            workbook.save(path)
        except Exception as config_save_exception:
            exception_message = "Below error is occurred while saving Config file in path {0} \n\t {1}".format(path,
                                                                                                               config_save_exception)
            raise exception_message

        return dict_config_main

    except Exception as config_file_read_error:
        exception_message = "Below error is occurred while reading Config file in path {0} \n\t {1}".format(path,
                                                                                                            config_file_read_error)
        raise exception_message